"""
Build branches index from official Excel source.

Usage:
  python -m app.data.build_branches_index
"""

from __future__ import annotations

import json
import logging
import re
from pathlib import Path
from typing import List, Optional

import openpyxl
from app.core.paths import EXCEL_BRANCHES_PATH

logger = logging.getLogger(__name__)

SOURCE_XLSX = EXCEL_BRANCHES_PATH
OUTPUT_JSON = Path(__file__).resolve().parent / "branches_index.json"
CITY_CANONICAL = [
    "الرياض",
    "جدة",
    "مكة",
    "الطائف",
    "الشرقية",
    "حائل",
    "القصيم",
    "وادي الدواسر",
    "السليل",
    "شقراء",
    "الزلفي",
    "المجمعة",
    "حفر الباطن",
    "الدوادمي",
    "ساجر",
    "عفيف",
    "الخرج",
    "الدلم",
    "حوطة بني تميم",
    "المزاحمية",
    "القويعية",
    "رماح",
    "ثادق",
]


def _clean_space(text: str) -> str:
    return re.sub(r"\s+", " ", (text or "").replace("\n", " ")).strip()


def _extract_phone(text: str) -> Optional[str]:
    if not text:
        return None
    m = re.search(r"\b(\d{8,12})\b", text)
    return m.group(1) if m else None


def _normalize_match(text: str) -> str:
    value = _clean_space(text)
    value = re.sub(r"[\u064B-\u065F\u0670\u0640]", "", value)
    value = value.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    value = value.replace("ى", "ي").replace("ؤ", "و").replace("ئ", "ي")
    value = re.sub(r"[^\w\s\u0600-\u06FF]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _extract_group_from_known_cities(raw: str) -> str:
    normalized = _normalize_match(raw)
    hits: list[tuple[int, str]] = []
    for city in CITY_CANONICAL:
        city_n = _normalize_match(city)
        idx = normalized.find(city_n)
        if idx >= 0:
            hits.append((idx, city))
    if not hits:
        return ""
    hits.sort(key=lambda x: x[0])
    ordered: list[str] = []
    seen = set()
    for _, city in hits:
        if city not in seen:
            seen.add(city)
            ordered.append(city)
    if len(ordered) == 1:
        return ordered[0]
    return " - ".join(ordered)


def _clean_group_title(raw: str) -> str:
    known = _extract_group_from_known_cities(raw)
    if known:
        return known
    text = _clean_space(raw)
    text = re.sub(r"\b\d{8,12}\b", " ", text)
    boilerplate = [
        "مختبرات وريد الطبية",
        "مختبرات وريد",
        "مواقع فروع",
        "موقع فرع",
        "موقع",
        "فروع منطقة",
        "فروع",
        "فرع",
        "للإتصال",
        "للاتصال",
        "للاتصال",
        "جودة حياتكم وريدنا",
        "جــودة حـيـاتـكـم وريــدنـا",
        "زورنا للمشاركة بدون شروط",
        "الفرع و ساعات العمل",
        "الافتتاح قريبا",
        "تم الافتتاح",
    ]
    for token in boilerplate:
        text = text.replace(token, " ")
    text = re.sub(r"[|,:]+", " ", text)
    text = re.sub(r"\s*-\s*", " - ", text)
    text = re.sub(r"^في\s+", "", text)
    text = _clean_space(text).strip("- ").strip()
    if not text:
        return "غير محدد"
    return text


def _clean_branch_name(raw: str) -> str:
    text = _clean_space(raw)
    if "فرع" in text:
        text = text[text.find("فرع") :]
    text = re.sub(r"الفرع و ساعات العمل", " ", text)
    text = _clean_space(text)
    return text


def _extract_hours(branch_name: str) -> str:
    if re.search(r"(٢٤|24)\s*ساعة", branch_name):
        return "٢٤ ساعة"
    return ""


def build_branches_index(source_path: Path = SOURCE_XLSX, output_path: Path = OUTPUT_JSON) -> List[dict]:
    if not source_path.exists():
        raise FileNotFoundError(f"Branches Excel not found: {source_path}")

    wb = openpyxl.load_workbook(source_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows: List[dict] = []
    current_group = "غير محدد"
    current_phone = ""

    for r in range(1, ws.max_row + 1):
        cell_a = ws.cell(r, 1).value
        cell_b = ws.cell(r, 2).value

        raw_a = _clean_space(str(cell_a)) if cell_a is not None else ""
        raw_b = _clean_space(str(cell_b)) if cell_b is not None else ""
        norm_a = _normalize_match(raw_a)

        is_map_row = raw_b.startswith("http://") or raw_b.startswith("https://")
        if is_map_row:
            branch_name = _clean_branch_name(raw_a)
            if not branch_name:
                continue
            hours = _extract_hours(branch_name)
            rows.append(
                {
                    "group": current_group,
                    "branch_name": branch_name,
                    "hours": hours,
                    "maps_url": raw_b,
                    "phone": current_phone or "",
                }
            )
            continue

        # Group/header rows
        if raw_a and ("فرع" in norm_a or "فروع" in norm_a or "مختبرات وريد" in norm_a or "مواقع" in norm_a):
            if "الفرع و ساعات العمل" in norm_a:
                continue
            current_group = _clean_group_title(raw_a)
            maybe_phone = _extract_phone(raw_a) or _extract_phone(raw_b)
            current_phone = maybe_phone or ""

    with output_path.open("w", encoding="utf-8") as f:
        json.dump(rows, f, ensure_ascii=False, indent=2)

    return rows


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
    data = build_branches_index()
    logger.info("Branches indexed: %s", len(data))
    logger.info("Output: %s", OUTPUT_JSON)


if __name__ == "__main__":
    main()
