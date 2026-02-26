"""
Build deterministic packages index from PAKAGE1.xlsx.

Usage:
  python -m app.data.build_packages_index
"""

from __future__ import annotations

import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import openpyxl

SOURCE_XLSX = Path(__file__).resolve().parent / "PAKAGE1.xlsx"
OUTPUT_JSON = Path(__file__).resolve().parent / "packages_index.json"
SHEET_NAME = "Sheet1"

HEADER_ROW = 2

ARABIC_DIACRITICS_RE = re.compile(r"[\u064B-\u065F\u0670]")
PUNCT_RE = re.compile(r"[^\w\s\u0600-\u06FF]")
WHITESPACE_RE = re.compile(r"\s+")

PRICE_RE = re.compile(r"([0-9\u0660-\u0669]+)\s*ريال")
ENGLISH_TOKEN_RE = re.compile(r"\b[A-Za-z][A-Za-z0-9/-]*\b")

PREFIX_RE = re.compile(r"^\s*(?:باقة|تحليل|تحاليل|تحالیل|فحص)\s+", re.IGNORECASE)

INCLUDES_KEYWORDS = (
    "تشمل",
    "تحتوي",
    "التحاليل المدرجة",
    "تفاصيل التحاليل",
    "المجالات التي يغطيها",
    "ما يغطيه التحليل",
    "الباقة تشمل",
)
TURNAROUND_KEYWORDS = ("مدة النتائج", "مدة التحليل")
SAMPLE_KEYWORDS = ("نوع العينة", "عينة", "دم", "براز")
AUDIENCE_KEYWORDS = ("للرجال", "للمرأة", "الأطفال")


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u00a0", " ").strip()


def _collapse_whitespace(text: str) -> str:
    return WHITESPACE_RE.sub(" ", text).strip()


def _normalize_text(text: str) -> str:
    value = _clean_text(text)
    value = ARABIC_DIACRITICS_RE.sub("", value)
    value = value.replace("ـ", "")
    value = value.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    value = value.replace("ى", "ي")
    value = value.replace("ؤ", "و")
    value = value.replace("ئ", "ي")
    value = value.lower()
    value = PUNCT_RE.sub(" ", value)
    return _collapse_whitespace(value)


def _to_ascii_digits(text: str) -> str:
    return text.translate(str.maketrans("٠١٢٣٤٥٦٧٨٩", "0123456789"))


def _slugify(value: str) -> str:
    base = _normalize_text(value).replace(" ", "-")
    base = re.sub(r"-+", "-", base).strip("-")
    return base or "item"


def _extract_by_keywords(text: str, keywords: tuple[str, ...]) -> str | None:
    if not text:
        return None
    lines = [line.strip() for line in str(text).splitlines() if line.strip()]
    for line in lines:
        if any(keyword in line for keyword in keywords):
            return line
    return None


def _parse_price(price_raw: str | None) -> tuple[int | None, str | None, str | None]:
    if not price_raw:
        return None, None, None

    raw = _clean_text(price_raw)
    if not raw:
        return None, None, None

    matches = list(PRICE_RE.finditer(raw))
    if not matches:
        return None, ("ريال" if "ريال" in raw else None), (_collapse_whitespace(raw) or None)

    price_match = matches[-1]
    digit_text = _to_ascii_digits(price_match.group(1))
    try:
        value = int(digit_text)
    except ValueError:
        value = None

    currency = "ريال"

    note = (raw[: price_match.start()] + " " + raw[price_match.end() :]).strip()
    note = re.sub(r"^\s*(?:ب|بـ)\s*", "", note)
    note = re.sub(r"\s*(?:ب|بـ)\s*$", "", note)
    note = _collapse_whitespace(note)
    note = note or None

    return value, currency, note


def _build_aliases(name_raw: str) -> list[str]:
    aliases: list[str] = []
    seen: set[str] = set()

    def add(alias: str) -> None:
        candidate = _collapse_whitespace(alias)
        if not candidate:
            return
        normalized = _normalize_text(candidate)
        if not normalized or normalized in seen:
            return
        seen.add(normalized)
        aliases.append(candidate)

    add(name_raw)

    stripped = PREFIX_RE.sub("", name_raw).strip()
    if stripped and _normalize_text(stripped) != _normalize_text(name_raw):
        add(stripped)

    for token in ENGLISH_TOKEN_RE.findall(name_raw):
        add(token)

    return aliases


def _is_empty(value: Any) -> bool:
    return _clean_text(value) == ""


def build_packages_index(
    source_path: Path = SOURCE_XLSX,
    output_path: Path = OUTPUT_JSON,
) -> list[dict[str, Any]]:
    if not source_path.exists():
        raise FileNotFoundError(f"Packages Excel not found: {source_path}")

    wb = openpyxl.load_workbook(source_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found in workbook")
    ws = wb[SHEET_NAME]

    records: list[dict[str, Any]] = []
    current_section = ""

    total_rows_scanned = ws.max_row
    category_rows_count = 0

    for row in range(1, ws.max_row + 1):
        if row == HEADER_ROW:
            continue

        a_raw = _clean_text(ws.cell(row, 1).value)
        b_raw = _clean_text(ws.cell(row, 2).value)
        c_raw = _clean_text(ws.cell(row, 3).value)

        # Category rows: A has text, B and C empty.
        if a_raw and not b_raw and not c_raw:
            current_section = a_raw
            category_rows_count += 1
            continue

        # Item rows: A and B non-empty, C optional.
        if not (a_raw and b_raw):
            continue

        name_norm = _normalize_text(a_raw)
        price_raw = c_raw or None
        price_value, price_currency, price_note = _parse_price(price_raw)

        record: dict[str, Any] = {
            "id": f"{_slugify(a_raw)}-r{row}",
            "sheet": SHEET_NAME,
            "row": row,
            "section": current_section or None,
            "name_raw": a_raw,
            "name_norm": name_norm,
            "description_raw": b_raw,
            "price_raw": price_raw,
            "price_value": price_value,
            "price_currency": price_currency,
            "price_note": price_note,
            "includes_text": _extract_by_keywords(b_raw, INCLUDES_KEYWORDS),
            "sample_type_text": _extract_by_keywords(b_raw, SAMPLE_KEYWORDS),
            "turnaround_text": _extract_by_keywords(b_raw, TURNAROUND_KEYWORDS),
            "audience_text": _extract_by_keywords((a_raw + "\n" + b_raw), AUDIENCE_KEYWORDS),
            "aliases": _build_aliases(a_raw),
            "duplicate_flag": False,
        }
        records.append(record)

    # Keep duplicates as separate records, but mark them.
    name_counts = Counter(record["name_norm"] for record in records)
    for record in records:
        if name_counts[record["name_norm"]] > 1:
            record["duplicate_flag"] = True

    with output_path.open("w", encoding="utf-8") as f:
        json.dump(records, f, ensure_ascii=False, indent=2)

    priced = sum(1 for r in records if r["price_raw"])
    missing = len(records) - priced
    duplicate_groups = sum(1 for count in name_counts.values() if count > 1)

    print(f"total rows scanned: {total_rows_scanned}")
    print(f"total category rows: {category_rows_count}")
    print(f"total package/test records: {len(records)}")
    print(f"records with price: {priced}")
    print(f"records missing price: {missing}")
    print(f"duplicates count: {duplicate_groups}")

    return records


def main() -> None:
    build_packages_index()


if __name__ == "__main__":
    main()
