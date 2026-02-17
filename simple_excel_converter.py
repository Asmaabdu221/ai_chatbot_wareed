"""
Simple Excel to JSON Converter using openpyxl
"""

import openpyxl
import json
from pathlib import Path

def excel_to_json(excel_path, output_path):
    """Convert Excel to JSON"""
    try:
        # Load workbook
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        result = {}
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                if cell.value:
                    headers.append(str(cell.value))
            
            # Get data rows
            data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                # Skip empty rows
                if not any(row):
                    continue
                
                # Create record
                record = {}
                for i, value in enumerate(row):
                    if i < len(headers):
                        record[headers[i]] = value
                
                data.append(record)
            
            result[sheet_name] = {
                "total_records": len(data),
                "columns": headers,
                "data": data
            }
        
        # Save JSON
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(result, f, ensure_ascii=False, indent=2, default=str)
        
        return result
    
    except Exception as e:
        print(f"Error: {e}")
        return None


if __name__ == "__main__":
    excel_file = r"c:\Users\asmaa\Downloads\testing.xlsx"
    output_file = r"c:\Users\asmaa\OneDrive\Documents\work\ai_chatbot_wareed\app\data\excel_data.json"
    
    result = excel_to_json(excel_file, output_file)
    
    if result:
        print(f"Success! Created: {output_file}")
        for sheet, info in result.items():
            print(f"  - {sheet}: {info['total_records']} records")
    else:
        print("Failed!")
