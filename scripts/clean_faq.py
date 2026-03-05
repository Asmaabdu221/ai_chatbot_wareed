from __future__ import annotations

import argparse
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_SOURCE_PATH = Path("app/data/sources/excel/faq.xlsx")
DEFAULT_SHEET_NAME = "Sheet1"
DEFAULT_OUTPUT_PATH = Path("app/data/runtime/lookup/faq_index.json")
DEFAULT_RAG_CLEAN_OUTPUT_PATH = Path("app/data/runtime/rag/faq_clean.jsonl")
DEFAULT_RAG_CHUNKS_OUTPUT_PATH = Path("app/data/runtime/rag/faq_chunks.jsonl")
DEFAULT_QA_REPORT_OUTPUT_PATH = Path("app/data/runtime/reports/faq_qa.json")

ARABIC_DIACRITICS_RE = re.compile(r"[\u064B-\u065F\u0670\u0640]")
SPACE_RE = re.compile(r"[ \t\x0b\x0c\r]+")
PUNCT_TO_REMOVE_RE = re.compile(r"[؟!،\.\:;\"'()\[\]\{\}\-_/\\]")


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value)


def _normalize_spaces(text: str) -> str:
    return SPACE_RE.sub(" ", text).strip()


def _normalize_spaces_preserve_newlines(text: str) -> str:
    lines = text.split("\n")
    cleaned = [_normalize_spaces(line) for line in lines]
    return "\n".join(cleaned).strip()


def normalize_question(text: str) -> str:
    value = _normalize_spaces(text or "")
    value = ARABIC_DIACRITICS_RE.sub("", value).replace("ـ", "")
    value = (
        value.replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace("ى", "ي")
        .replace("ة", "ه")
    )
    value = PUNCT_TO_REMOVE_RE.sub(" ", value)
    value = "".join(ch for ch in value if not unicodedata.category(ch).startswith("P"))
    value = _normalize_spaces(value)
    return value.lower()


def _is_header_row(col_a: str, col_b: str) -> bool:
    return "السؤال" in col_a and "الإجابة" in col_b


def parse_faq_rows(
    source_path: Path, sheet_name: str
) -> tuple[list[dict[str, str]], dict[str, int]]:
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")

    sheet = workbook[sheet_name]
    cleaned_records: list[dict[str, str]] = []
    seen_questions: set[str] = set()
    stats = {
        "total_rows_scanned": 0,
        "valid_qa_pairs_written": 0,
        "deduped_count": 0,
        "empty_dropped_count": 0,
    }

    for a_raw, b_raw in sheet.iter_rows(min_col=1, max_col=2, values_only=True):
        stats["total_rows_scanned"] += 1
        col_a_raw = _to_text(a_raw)
        col_b_raw = _to_text(b_raw)

        col_a = _normalize_spaces(col_a_raw)
        col_b = _normalize_spaces_preserve_newlines(col_b_raw)

        # Rule 1: skip header rows.
        if _is_header_row(col_a, col_b):
            continue

        # Rule 2: drop empty rows.
        if not col_a and not col_b:
            stats["empty_dropped_count"] += 1
            continue

        # Only keep complete Q/A rows.
        if not col_a or not col_b:
            stats["empty_dropped_count"] += 1
            continue

        # Rule 6: deduplicate by normalized question.
        q_key = normalize_question(col_a)
        if not q_key or q_key in seen_questions:
            stats["deduped_count"] += 1
            continue
        seen_questions.add(q_key)

        cleaned_records.append({"question": col_a, "q_norm": q_key, "answer": col_b})
        stats["valid_qa_pairs_written"] += 1

    return cleaned_records, stats


def build_lookup_index(cleaned_records: list[dict[str, str]]) -> list[dict[str, Any]]:
    index: list[dict[str, Any]] = []
    for idx, record in enumerate(cleaned_records, start=1):
        index.append(
            {
                "id": f"faq::{idx}",
                "q": record["question"],
                "q_norm": record["q_norm"],
                "a": record["answer"],
                "tags": ["faq"],
            }
        )
    return index


def write_json_array(items: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def write_json_object(item: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(item, f, ensure_ascii=False, indent=2)


def write_jsonl(items: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        for item in items:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def build_rag_clean_records(lookup_index: list[dict[str, Any]]) -> list[dict[str, str]]:
    records: list[dict[str, str]] = []
    for item in lookup_index:
        records.append(
            {
                "source": "faq",
                "id": item["id"],
                "question": item["q"],
                "answer": item["a"],
                "q_norm": item["q_norm"],
            }
        )
    return records


def build_rag_chunks(lookup_index: list[dict[str, Any]]) -> list[dict[str, Any]]:
    chunks: list[dict[str, Any]] = []
    for item in lookup_index:
        chunks.append(
            {
                "id": item["id"],
                "text": f"سؤال: {item['q']}\nإجابة: {item['a']}",
                "metadata": {"source": "faq", "type": "faq", "q_norm": item["q_norm"]},
            }
        )
    return chunks


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Clean FAQ sheet rows and preview output")
    parser.add_argument(
        "--source",
        type=Path,
        default=DEFAULT_SOURCE_PATH,
        help=f"Path to FAQ Excel file (default: {DEFAULT_SOURCE_PATH})",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"Worksheet name (default: {DEFAULT_SHEET_NAME})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Path to FAQ lookup output JSON (default: {DEFAULT_OUTPUT_PATH})",
    )
    return parser


def main() -> None:
    args = _build_arg_parser().parse_args()
    records, stats = parse_faq_rows(args.source, args.sheet)
    lookup_index = build_lookup_index(records)
    write_json_array(lookup_index, args.output)
    write_jsonl(build_rag_clean_records(lookup_index), DEFAULT_RAG_CLEAN_OUTPUT_PATH)
    write_jsonl(build_rag_chunks(lookup_index), DEFAULT_RAG_CHUNKS_OUTPUT_PATH)
    qa_report = {
        "input_file": args.source.as_posix(),
        "lookup_output": args.output.as_posix(),
        "rag_clean_output": DEFAULT_RAG_CLEAN_OUTPUT_PATH.as_posix(),
        "rag_chunks_output": DEFAULT_RAG_CHUNKS_OUTPUT_PATH.as_posix(),
        "total_rows_scanned": stats["total_rows_scanned"],
        "valid_written": stats["valid_qa_pairs_written"],
        "deduped": stats["deduped_count"],
        "dropped_empty": stats["empty_dropped_count"],
        "sample_items": lookup_index[:3],
    }
    write_json_object(qa_report, DEFAULT_QA_REPORT_OUTPUT_PATH)

    print(f"total rows scanned: {stats['total_rows_scanned']}")
    print(f"valid qa pairs written: {stats['valid_qa_pairs_written']}")
    print(f"deduped count: {stats['deduped_count']}")
    print(f"empty dropped count: {stats['empty_dropped_count']}")
    print(f"lookup output: {args.output.as_posix()}")
    print(f"rag clean output: {DEFAULT_RAG_CLEAN_OUTPUT_PATH.as_posix()}")
    print(f"rag chunks output: {DEFAULT_RAG_CHUNKS_OUTPUT_PATH.as_posix()}")
    print(f"qa report output: {DEFAULT_QA_REPORT_OUTPUT_PATH.as_posix()}")
    for item in lookup_index[:3]:
        print(f"{item['q']} -> {item['q_norm']}")


if __name__ == "__main__":
    main()
