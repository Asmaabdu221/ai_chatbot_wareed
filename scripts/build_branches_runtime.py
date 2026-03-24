from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl

INPUT_XLSX_PATH = Path("app/data/sources/excel/branches_structured_review_updated.xlsx")
INPUT_SHEET_NAME = "Structured_Branches"
OUTPUT_JSONL_PATH = Path("app/data/runtime/rag/branches_with_coordinates.jsonl")

_CITY_ALIASES = {
    "الرياض": "الرياض",
    "رياض": "الرياض",
    "جدة": "جدة",
    "جده": "جدة",
    "مكة": "مكة المكرمة",
    "مكه": "مكة المكرمة",
    "مكة المكرمة": "مكة المكرمة",
    "مكه المكرمه": "مكة المكرمة",
    "المدينة": "المدينة المنورة",
    "المدينه": "المدينة المنورة",
    "المدينة المنورة": "المدينة المنورة",
    "المدينه المنوره": "المدينة المنورة",
    "الطائف": "الطائف",
    "الطايف": "الطائف",
}


def _safe_str(value: Any) -> str:
    return str(value or "").strip()


def _normalize_arabic_for_key(text: str) -> str:
    value = _safe_str(text).lower()
    value = (
        value.replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace("ى", "ي")
        .replace("ة", "ه")
        .replace("ؤ", "و")
        .replace("ئ", "ي")
    )
    value = re.sub(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]", "", value)
    value = value.replace("ـ", "")
    value = re.sub(r"\s+", " ", value).strip()
    return value


_CITY_ALIASES_NORM = {
    _normalize_arabic_for_key(k): v for k, v in _CITY_ALIASES.items()
}


def _canonical_city(city: str) -> str:
    raw = _safe_str(city)
    if not raw:
        return ""
    key = _normalize_arabic_for_key(raw)
    return _CITY_ALIASES_NORM.get(key, raw)


def _to_float_or_none(value: Any) -> float | None:
    text = _safe_str(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_bool(value: Any, default: bool = True) -> bool:
    if isinstance(value, bool):
        return value
    text = _safe_str(value).lower()
    if not text:
        return default
    if text in {"true", "1", "yes"}:
        return True
    if text in {"false", "0", "no"}:
        return False
    return default


def _slug_id_part(text: str) -> str:
    value = _normalize_arabic_for_key(text)
    value = re.sub(r"[^a-z0-9\u0600-\u06FF\s_-]", "", value)
    value = re.sub(r"\s+", "_", value).strip("_")
    return value or "unknown"


def _fallback_id(city: str, branch_name: str) -> str:
    return f"branch::{_slug_id_part(city)}_{_slug_id_part(branch_name)}"


def _header_index(header_row: list[Any]) -> dict[str, int]:
    idx: dict[str, int] = {}
    for i, name in enumerate(header_row):
        key = _safe_str(name).lower()
        if key:
            idx[key] = i
    return idx


def _cell(row: list[Any], idx: dict[str, int], key: str) -> Any:
    pos = idx.get(key.lower())
    if pos is None or pos >= len(row):
        return ""
    return row[pos]


def build_runtime_branches(
    input_path: Path = INPUT_XLSX_PATH,
    output_path: Path = OUTPUT_JSONL_PATH,
) -> dict[str, int]:
    if not input_path.exists():
        raise FileNotFoundError(f"Input not found: {input_path}")

    wb = openpyxl.load_workbook(input_path, data_only=True)
    if INPUT_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{INPUT_SHEET_NAME}' in {input_path}")

    ws = wb[INPUT_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{INPUT_SHEET_NAME}' is empty")

    header = _header_index(list(rows[0]))
    required = [
        "city",
        "district",
        "branch name",
        "hours",
        "map url",
        "maps url",
        "latitude",
        "longitude",
        "contact phone",
        "id",
        "is active",
        "source raw text",
        "section",
    ]
    missing = [k for k in required if k not in header]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    output: list[dict[str, Any]] = []
    for raw_row in rows[1:]:
        row = list(raw_row)
        branch_name = _safe_str(_cell(row, header, "branch name"))
        city = _canonical_city(_safe_str(_cell(row, header, "city")))
        district = _safe_str(_cell(row, header, "district"))
        if not branch_name or not city:
            continue

        map_url = _safe_str(_cell(row, header, "map url"))
        maps_url = _safe_str(_cell(row, header, "maps url")) or map_url
        map_url = map_url or maps_url

        record = {
            "source": "branches",
            "id": _safe_str(_cell(row, header, "id")) or _fallback_id(city, branch_name),
            "city": city,
            "district": district,
            "branch_name": branch_name,
            "hours": _safe_str(_cell(row, header, "hours")) or None,
            "map_url": map_url,
            "maps_url": maps_url,
            "latitude": _to_float_or_none(_cell(row, header, "latitude")),
            "longitude": _to_float_or_none(_cell(row, header, "longitude")),
            "contact_phone": _safe_str(_cell(row, header, "contact phone")),
            "is_active": _to_bool(_cell(row, header, "is active"), default=True),
            "raw_text": _safe_str(_cell(row, header, "source raw text")),
            "section": _safe_str(_cell(row, header, "section")),
        }
        output.append(record)

    output.sort(key=lambda r: (_safe_str(r.get("city")), _safe_str(r.get("branch_name"))))

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        for item in output:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")

    return {"written": len(output)}


if __name__ == "__main__":
    stats = build_runtime_branches()
    print(f"INPUT : {INPUT_XLSX_PATH.as_posix()}")
    print(f"SHEET : {INPUT_SHEET_NAME}")
    print(f"OUTPUT: {OUTPUT_JSONL_PATH.as_posix()}")
    print(f"ROWS  : {stats['written']}")
