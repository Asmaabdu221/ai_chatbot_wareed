from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl

INPUT_XLSX_PATH = Path("app/data/sources/excel/analyses_business_structured_review.xlsx")
INPUT_SHEET_NAME = "Structured_Analyses"
OUTPUT_JSONL_PATH = Path("app/data/runtime/rag/tests_business_clean.jsonl")


def _safe_str(value: Any) -> str:
    return str(value or "").strip()


def _normalize_key(text: str) -> str:
    return re.sub(r"\s+", " ", _safe_str(text).lower()).strip()


def _to_int_or_none(value: Any) -> int | None:
    text = _safe_str(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _to_float_or_none(value: Any) -> float | None:
    text = _safe_str(value)
    if not text:
        return None
    cleaned = re.sub(r"[^\d.\-]", "", text)
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def _to_bool(value: Any, *, default: bool = True) -> bool:
    if isinstance(value, bool):
        return value
    text = _safe_str(value).lower()
    if not text:
        return default
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return default


def _split_values(value: Any) -> list[str]:
    text = _safe_str(value)
    if not text:
        return []
    parts = re.split(r"[\n,؛;|]+", text)
    out = [part.strip(" -\t") for part in parts if part and part.strip(" -\t")]
    # Preserve order while deduplicating.
    seen: set[str] = set()
    unique: list[str] = []
    for part in out:
        if part in seen:
            continue
        seen.add(part)
        unique.append(part)
    return unique


def _split_alias_terms(*values: Any) -> list[str]:
    merged: list[str] = []
    for value in values:
        merged.extend(_split_values(value))
    seen: set[str] = set()
    unique: list[str] = []
    for part in merged:
        if part in seen:
            continue
        seen.add(part)
        unique.append(part)
    return unique


def _normalize_arabic_simple(text: str) -> str:
    value = _safe_str(text).lower()
    if not value:
        return ""
    value = (
        value.replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace("ى", "ي")
        .replace("ة", "ه")
    )
    value = re.sub(r"[\u0610-\u061A\u064B-\u065F\u0670\u06D6-\u06ED]", "", value)
    value = value.replace("ـ", "")
    value = re.sub(r"[^A-Za-z0-9\u0660-\u0669\u06F0-\u06F9\u0621-\u063A\u0641-\u064A\s]", " ", value)
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _header_index(header_row: list[Any]) -> dict[str, int]:
    index: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = _normalize_key(cell)
        if key:
            index[key] = i
    return index


def _cell(row: list[Any], index: dict[str, int], key: str) -> Any:
    pos = index.get(_normalize_key(key))
    if pos is None or pos >= len(row):
        return ""
    return row[pos]


def _slug(text: str) -> str:
    value = _safe_str(text)
    value = re.sub(r"\s+", "_", value).strip("_")
    value = re.sub(r"[^a-zA-Z0-9_\u0600-\u06FF]", "", value)
    return value or "analysis"


def _fallback_id(source_row: int | None, test_name_ar: str) -> str:
    row_part = str(source_row or "unknown")
    return f"analysis::{row_part}_{_slug(test_name_ar)}"


def build_tests_business_runtime(
    input_path: Path = INPUT_XLSX_PATH,
    output_path: Path = OUTPUT_JSONL_PATH,
) -> dict[str, int]:
    if not input_path.exists():
        raise FileNotFoundError(f"Input not found: {input_path.as_posix()}")

    wb = openpyxl.load_workbook(input_path, data_only=True)
    if INPUT_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{INPUT_SHEET_NAME}' in {input_path.as_posix()}")

    ws = wb[INPUT_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{INPUT_SHEET_NAME}' is empty")

    header = _header_index(list(rows[0]))
    required_columns = [
        "row",
        "analysis name ar",
        "category",
        "price",
        "preparation",
        "symptoms",
        "complementary tests",
        "alternative tests",
        "sample type",
        "english name",
        "code / alt name",
        "matched name",
        "match score",
        "primary use cases",
        "review issues",
    ]
    missing = [col for col in required_columns if _normalize_key(col) not in header]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    output_rows: list[dict[str, Any]] = []
    for excel_row in rows[1:]:
        row = list(excel_row)
        test_name_ar = _safe_str(_cell(row, header, "analysis name ar"))
        if not test_name_ar:
            continue

        source_row = _to_int_or_none(_cell(row, header, "row"))
        english_name = _safe_str(_cell(row, header, "english name"))
        code_alt_name = _safe_str(_cell(row, header, "code / alt name"))
        matched_name = _safe_str(_cell(row, header, "matched name"))
        alias_terms = _split_alias_terms(english_name, code_alt_name, matched_name)
        match_terms = [test_name_ar] + alias_terms
        match_terms_norm = [_normalize_arabic_simple(x) for x in match_terms if _normalize_arabic_simple(x)]
        price_raw = _safe_str(_cell(row, header, "price"))
        record = {
            "source": "tests_business",
            "id": _safe_str(_cell(row, header, "id")) or _fallback_id(source_row, test_name_ar),
            "test_name_ar": test_name_ar,
            "english_name": english_name,
            "code_alt_name": code_alt_name,
            "matched_name": matched_name,
            "category": _safe_str(_cell(row, header, "category")),
            "benefit": _safe_str(_cell(row, header, "primary use cases")),
            "price_raw": price_raw,
            "price_number": _to_float_or_none(price_raw),
            "sample_type": _safe_str(_cell(row, header, "sample type")),
            "symptoms": _split_values(_cell(row, header, "symptoms")),
            "preparation": _safe_str(_cell(row, header, "preparation")),
            "complementary_tests": _split_values(_cell(row, header, "complementary tests")),
            "alternative_tests": _split_values(_cell(row, header, "alternative tests")),
            "alias_terms": alias_terms,
            "match_terms": match_terms,
            "match_terms_norm": match_terms_norm,
            "match_score": _to_float_or_none(_cell(row, header, "match score")),
            "review_issues": _safe_str(_cell(row, header, "review issues")),
            "source_row": source_row,
            "is_active": _to_bool(_cell(row, header, "is active"), default=True),
        }
        output_rows.append(record)

    output_rows.sort(
        key=lambda r: (
            int(r.get("source_row") or 10**9),
            _safe_str(r.get("test_name_ar")),
            _safe_str(r.get("id")),
        )
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        for record in output_rows:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    return {"written": len(output_rows)}


if __name__ == "__main__":
    stats = build_tests_business_runtime()
    print(f"INPUT : {INPUT_XLSX_PATH.as_posix()}")
    print(f"SHEET : {INPUT_SHEET_NAME}")
    print(f"OUTPUT: {OUTPUT_JSONL_PATH.as_posix()}")
    print(f"ROWS  : {stats['written']}")
