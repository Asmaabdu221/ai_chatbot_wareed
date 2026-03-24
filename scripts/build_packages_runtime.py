from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl

INPUT_XLSX_PATH = Path("app/data/sources/excel/packages_structured_review.xlsx")
INPUT_SHEET_NAME = "Structured_Packages"
OUTPUT_JSONL_PATH = Path("app/data/runtime/rag/packages_clean.jsonl")


def _safe_str(value: Any) -> str:
    return str(value or "").strip()


def _to_int_or_none(value: Any) -> int | None:
    text = _safe_str(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _to_float_or_none(value: Any) -> float | None:
    text = _safe_str(value)
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def _to_bool(value: Any, default: bool = True) -> bool:
    if isinstance(value, bool):
        return value
    text = _safe_str(value).lower()
    if not text:
        return default
    if text in {"yes", "true", "1", "y"}:
        return True
    if text in {"no", "false", "0", "n"}:
        return False
    return default


def _normalize_key(text: str) -> str:
    value = _safe_str(text).lower()
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _header_index(header_row: list[Any]) -> dict[str, int]:
    index: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = _normalize_key(_safe_str(cell))
        if key:
            index[key] = i
    return index


def _cell(row: list[Any], index: dict[str, int], key: str) -> Any:
    pos = index.get(_normalize_key(key))
    if pos is None or pos >= len(row):
        return ""
    return row[pos]


def _fallback_id(source_row: int | None, package_name: str) -> str:
    source_part = str(source_row or "unknown")
    name_part = re.sub(r"\s+", "_", _safe_str(package_name)).strip("_")
    name_part = re.sub(r"[^a-zA-Z0-9_\u0600-\u06FF]", "", name_part) or "package"
    return f"package::{source_part}_{name_part}"


def build_packages_runtime(
    input_path: Path = INPUT_XLSX_PATH,
    output_path: Path = OUTPUT_JSONL_PATH,
) -> dict[str, int]:
    if not input_path.exists():
        raise FileNotFoundError(f"Input not found: {input_path.as_posix()}")

    wb = openpyxl.load_workbook(input_path, data_only=True)
    if INPUT_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{INPUT_SHEET_NAME}' in {input_path.as_posix()}")

    ws = wb[INPUT_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{INPUT_SHEET_NAME}' is empty")

    header = _header_index(list(rows[0]))
    required_columns = [
        "source row",
        "main category",
        "offering type",
        "package name",
        "description short",
        "description full",
        "price raw",
        "price number",
        "currency",
        "included count",
        "runtime present",
        "review issues",
    ]
    missing = [col for col in required_columns if _normalize_key(col) not in header]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    output_rows: list[dict[str, Any]] = []
    for excel_row in rows[1:]:
        row = list(excel_row)
        package_name = _safe_str(_cell(row, header, "package name"))
        offering_type = _safe_str(_cell(row, header, "offering type"))
        main_category = _safe_str(_cell(row, header, "main category"))
        source_row = _to_int_or_none(_cell(row, header, "source row"))

        if not package_name:
            continue

        record = {
            "source": "packages",
            "id": _safe_str(_cell(row, header, "id")) or _fallback_id(source_row, package_name),
            "main_category": main_category,
            "offering_type": offering_type,
            "package_name": package_name,
            "description_short": _safe_str(_cell(row, header, "description short")),
            "description_full": _safe_str(_cell(row, header, "description full")),
            "price_raw": _safe_str(_cell(row, header, "price raw")),
            "price_number": _to_float_or_none(_cell(row, header, "price number")),
            "currency": _safe_str(_cell(row, header, "currency")),
            "included_count": _to_int_or_none(_cell(row, header, "included count")),
            "runtime_present": _to_bool(_cell(row, header, "runtime present"), default=False),
            "review_issues": _safe_str(_cell(row, header, "review issues")),
            "source_row": source_row,
            "is_active": True,
        }
        output_rows.append(record)

    output_rows.sort(
        key=lambda r: (
            int(r.get("source_row") or 10**9),
            _safe_str(r.get("package_name")),
            _safe_str(r.get("id")),
        )
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        for record in output_rows:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    return {"written": len(output_rows)}


if __name__ == "__main__":
    stats = build_packages_runtime()
    print(f"INPUT : {INPUT_XLSX_PATH.as_posix()}")
    print(f"SHEET : {INPUT_SHEET_NAME}")
    print(f"OUTPUT: {OUTPUT_JSONL_PATH.as_posix()}")
    print(f"ROWS  : {stats['written']}")
