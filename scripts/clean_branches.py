from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

import openpyxl

DEFAULT_SOURCE_PATH = Path("app/data/sources/excel/branches.xlsx")
DEFAULT_SHEET_NAME = "Sheet1"
DEFAULT_OUTPUT_PATH = Path("app/data/runtime/rag/branches_clean.jsonl")
DEFAULT_CHUNKS_OUTPUT_PATH = Path("app/data/runtime/rag/branches_chunks.jsonl")
DEFAULT_QA_OUTPUT_PATH = Path("app/data/runtime/reports/branches_qa.json")
DEFAULT_CONTACT_PHONE = "920003694"

ARABIC_DIACRITICS_RE = re.compile(r"[\u064B-\u065F\u0670\u0640]")
SPACE_RE = re.compile(r"\s+")
SECTION_HINTS = ("مواقع", "موقع", "فروع", "فرع", "مختبرات وريد")
HOURS_RE = re.compile(r"(24|٢٤)\s*ساعة")
LONG_DIGITS_RE = re.compile(r"[0-9٠-٩]{6,}")


def _clean_cell(value: object) -> str:
    if value is None:
        return ""
    text = str(value).replace("\n", " ").strip()
    return SPACE_RE.sub(" ", text)


def _normalize_ar(text: str) -> str:
    value = ARABIC_DIACRITICS_RE.sub("", text or "")
    value = (
        value.replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace("ى", "ي")
        .replace("ؤ", "و")
        .replace("ئ", "ي")
        .replace("ة", "ه")
    )
    return SPACE_RE.sub(" ", value).strip().lower()


def _is_header_row(col_a: str, col_b: str) -> bool:
    a_norm = _normalize_ar(col_a)
    b_norm = _normalize_ar(col_b)
    return "الفرع" in a_norm and "رابط الموقع" in b_norm


def _is_section_row(col_a: str, col_b: str) -> bool:
    if not col_a or col_b:
        return False
    a_norm = _normalize_ar(col_a)
    return any(_normalize_ar(token) in a_norm for token in SECTION_HINTS)


def _is_maps_url(value: str) -> bool:
    return value.lower().startswith("http")


def _clean_section_title(text: str) -> str:
    value = _clean_cell(text)

    # Remove contact snippets and phone-like long digit sequences.
    value = re.sub(r"\bللاتصال\b\s*[:\-]?\s*(?:[0-9٠-٩][0-9٠-٩\-\s]{5,})", " ", value)
    value = re.sub(r"\bللاتصال\b", " ", value)
    value = LONG_DIGITS_RE.sub(" ", value)

    # Remove known boilerplate fragments found in section rows.
    boilerplate_patterns = [
        r"مختبرات\s+وريد(?:\s+الطبية)?",
        r"زورنا\s+للمشاركة\s+بدون\s+شروط",
        r"ج[ـ\s]*ودة\s+ح[ـ\s]*ي[ـ\s]*ات[ـ\s]*ك[ـ\s]*م\s+و[رر]ي[ـ\s]*د[ـ\s]*ن[ـ\s]*ا",
        r"الافتتاح\s+قريبا",
    ]
    for pattern in boilerplate_patterns:
        value = re.sub(pattern, " ", value, flags=re.IGNORECASE)

    value = re.sub(r"\bفروع\s+في\b", "فروع ", value)
    value = value.strip("-–—،,;: ")
    value = SPACE_RE.sub(" ", value).strip()
    return value


def _extract_hours_and_branch_name(text: str) -> tuple[str, str | None]:
    match = HOURS_RE.search(text)
    if not match:
        return text, None

    hours = _clean_cell(match.group(0))
    branch_name = _clean_cell(text[: match.start()]).strip("-–—،, ")
    return branch_name, hours


def parse_branches_rows(
    source_path: Path, sheet_name: str
) -> tuple[list[dict[str, Any]], dict[str, int], list[tuple[str, str]]]:
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {workbook.sheetnames}")

    sheet = workbook[sheet_name]
    records: list[dict[str, Any]] = []
    current_section = ""
    section_changes: list[tuple[str, str]] = []
    stats = {
        "total_rows_scanned": 0,
        "valid_branches_written": 0,
        "skipped_blank_rows": 0,
        "skipped_header_rows": 0,
        "skipped_invalid_urls": 0,
    }

    for _row_index, (a_raw, b_raw) in enumerate(
        sheet.iter_rows(min_col=1, max_col=2, values_only=True), start=1
    ):
        stats["total_rows_scanned"] += 1
        col_a = _clean_cell(a_raw)
        col_b = _clean_cell(b_raw)

        # Rule 1: drop fully empty rows.
        if not col_a and not col_b:
            stats["skipped_blank_rows"] += 1
            continue

        # Rule 2: skip repeated header rows.
        if _is_header_row(col_a, col_b):
            stats["skipped_header_rows"] += 1
            continue

        # Rule 3 + 4: section titles with empty col B update section context.
        if _is_section_row(col_a, col_b):
            cleaned_section = _clean_section_title(col_a)
            if len(section_changes) < 3:
                section_changes.append((col_a, cleaned_section))
            current_section = cleaned_section
            continue

        # Data rows require branch text and valid URL.
        if not col_a:
            continue
        if not _is_maps_url(col_b):
            stats["skipped_invalid_urls"] += 1
            continue

        branch_name, hours = _extract_hours_and_branch_name(col_a)
        if not branch_name:
            branch_name = col_a

        record = {
            "source": "branches",
            "section": current_section or None,
            "branch_name": branch_name,
            "hours": hours,
            "raw_text": col_a,
            "map_url": col_b,
            "contact_phone": DEFAULT_CONTACT_PHONE,
        }
        records.append(record)
        stats["valid_branches_written"] += 1

    return records, stats, section_changes


def write_jsonl(records: list[dict[str, Any]], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        for record in records:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")


def build_chunks(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    chunks: list[dict[str, Any]] = []
    for idx, record in enumerate(records, start=1):
        section = record.get("section")
        hours = record.get("hours")
        section_text = section if section else "غير محدد"
        hours_text = hours if hours else "غير مذكور"
        text = (
            f"الفرع: {record['branch_name']}. "
            f"القسم: {section_text}. "
            f"ساعات العمل: {hours_text}. "
            f"رابط الموقع: {record['map_url']}."
        )
        chunks.append(
            {
                "id": f"branches::{idx}",
                "text": text,
                "metadata": {
                    "source": "branches",
                    "section": section,
                    "branch_name": record["branch_name"],
                    "hours": hours,
                    "map_url": record["map_url"],
                    "contact_phone": DEFAULT_CONTACT_PHONE,
                },
            }
        )
    return chunks


def write_json(data: dict[str, Any], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def build_qa_report(
    source_path: Path,
    clean_output_path: Path,
    chunks_output_path: Path,
    records: list[dict[str, Any]],
    chunks: list[dict[str, Any]],
    stats: dict[str, int],
) -> dict[str, Any]:
    unique_sections: list[str] = []
    seen_sections: set[str] = set()
    missing_section_count = 0
    missing_hours_count = 0

    for record in records:
        section = record.get("section")
        hours = record.get("hours")
        if section:
            if section not in seen_sections:
                seen_sections.add(section)
                unique_sections.append(section)
        else:
            missing_section_count += 1
        if not hours:
            missing_hours_count += 1

    return {
        "input_file": source_path.as_posix(),
        "clean_output": clean_output_path.as_posix(),
        "chunks_output": chunks_output_path.as_posix(),
        "total_rows_scanned": stats["total_rows_scanned"],
        "valid_branches_written": stats["valid_branches_written"],
        "unique_sections_count": len(unique_sections),
        "unique_sections_sample": unique_sections[:10],
        "missing_section_count": missing_section_count,
        "missing_hours_count": missing_hours_count,
        "sample_clean_records": records[:3],
        "sample_chunks": chunks[:3],
    }


def _build_arg_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Row-based cleaner/parser for branches.xlsx")
    parser.add_argument(
        "--source",
        type=Path,
        default=DEFAULT_SOURCE_PATH,
        help=f"Path to branches Excel file (default: {DEFAULT_SOURCE_PATH})",
    )
    parser.add_argument(
        "--sheet",
        default=DEFAULT_SHEET_NAME,
        help=f"Worksheet name (default: {DEFAULT_SHEET_NAME})",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT_PATH,
        help=f"Path to output JSONL file (default: {DEFAULT_OUTPUT_PATH})",
    )
    return parser


def main() -> None:
    args = _build_arg_parser().parse_args()
    records, stats, section_changes = parse_branches_rows(args.source, args.sheet)
    write_jsonl(records, args.output)
    chunks = build_chunks(records)
    write_jsonl(chunks, DEFAULT_CHUNKS_OUTPUT_PATH)
    qa_report = build_qa_report(
        source_path=args.source,
        clean_output_path=args.output,
        chunks_output_path=DEFAULT_CHUNKS_OUTPUT_PATH,
        records=records,
        chunks=chunks,
        stats=stats,
    )
    write_json(qa_report, DEFAULT_QA_OUTPUT_PATH)

    print(f"Total rows scanned: {stats['total_rows_scanned']}")
    print(f"Valid branches written: {stats['valid_branches_written']}")
    print(f"Skipped blank rows: {stats['skipped_blank_rows']}")
    print(f"Skipped header rows: {stats['skipped_header_rows']}")
    print(f"Skipped invalid URLs: {stats['skipped_invalid_urls']}")
    print(f"Chunks written: {len(chunks)}")
    print(f"QA report path: {DEFAULT_QA_OUTPUT_PATH.as_posix()}")
    print(f"Unique sections count: {qa_report['unique_sections_count']}")
    print(f"Missing section count: {qa_report['missing_section_count']}")
    print(f"Missing hours count: {qa_report['missing_hours_count']}")
    for old_section, new_section in section_changes:
        print(f"Section sample: {old_section} -> {new_section}")


if __name__ == "__main__":
    main()
