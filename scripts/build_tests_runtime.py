from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Any

import openpyxl

INPUT_XLSX_PATH = Path("app/data/sources/excel/tests_structured_review.xlsx")
INPUT_SHEET_NAME = "Structured_Tests"
OUTPUT_JSONL_PATH = Path("app/data/runtime/rag/tests_clean.jsonl")


def _safe_str(value: Any) -> str:
    return str(value or "").strip()


def _to_int_or_none(value: Any) -> int | None:
    text = _safe_str(value)
    if not text:
        return None
    try:
        return int(float(text))
    except ValueError:
        return None


def _to_bool(value: Any, *, default: bool = False) -> bool:
    if isinstance(value, bool):
        return value
    text = _safe_str(value).lower()
    if not text:
        return default
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    return default


def _normalize_key(text: str) -> str:
    value = _safe_str(text).lower()
    value = re.sub(r"\s+", " ", value).strip()
    return value


def _header_index(header_row: list[Any]) -> dict[str, int]:
    index: dict[str, int] = {}
    for i, cell in enumerate(header_row):
        key = _normalize_key(_safe_str(cell))
        if key:
            index[key] = i
    return index


def _cell(row: list[Any], index: dict[str, int], key: str) -> Any:
    pos = index.get(_normalize_key(key))
    if pos is None or pos >= len(row):
        return ""
    return row[pos]


def _split_tokens(value: Any) -> list[str]:
    text = _safe_str(value)
    if not text:
        return []
    parts = [part.strip() for part in text.split(",")]
    return [part for part in parts if part]


def _fallback_id(title: str, source_row: int) -> str:
    title_part = re.sub(r"\s+", "_", _safe_str(title)).strip("_")
    title_part = re.sub(r"[^a-zA-Z0-9_\u0600-\u06FF]", "", title_part) or "test"
    return f"test::{source_row}_{title_part}"


def build_tests_runtime(
    input_path: Path = INPUT_XLSX_PATH,
    output_path: Path = OUTPUT_JSONL_PATH,
) -> dict[str, int]:
    if not input_path.exists():
        raise FileNotFoundError(f"Input not found: {input_path.as_posix()}")

    wb = openpyxl.load_workbook(input_path, data_only=True)
    if INPUT_SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Missing sheet '{INPUT_SHEET_NAME}' in {input_path.as_posix()}")

    ws = wb[INPUT_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError(f"Sheet '{INPUT_SHEET_NAME}' is empty")

    header = _header_index(list(rows[0]))
    required_columns = [
        "id",
        "page type",
        "source type",
        "domain",
        "test name (ar)",
        "title",
        "h1",
        "code tokens",
        "tags",
        "summary ar",
        "content clean",
        "url",
        "summary length",
        "content length",
        "chunk exists",
        "chunk count",
        "review issues",
    ]
    missing = [col for col in required_columns if _normalize_key(col) not in header]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    output_rows: list[dict[str, Any]] = []
    for row_idx, excel_row in enumerate(rows[1:], start=2):
        row = list(excel_row)
        title = _safe_str(_cell(row, header, "title"))
        test_name_ar = _safe_str(_cell(row, header, "test name (ar)"))
        record_id = _safe_str(_cell(row, header, "id")) or _fallback_id(title or test_name_ar, row_idx)

        if not (record_id and (title or test_name_ar)):
            continue

        content_clean = _safe_str(_cell(row, header, "content clean"))
        summary_ar = _safe_str(_cell(row, header, "summary ar"))
        summary_length = _to_int_or_none(_cell(row, header, "summary length"))
        content_length = _to_int_or_none(_cell(row, header, "content length"))

        record = {
            "source": "tests",
            "id": record_id,
            "page_type": _safe_str(_cell(row, header, "page type")),
            "source_type": _safe_str(_cell(row, header, "source type")),
            "domain": _safe_str(_cell(row, header, "domain")),
            "test_name_ar": test_name_ar,
            "title": title,
            "h1": _safe_str(_cell(row, header, "h1")),
            "code_tokens": _split_tokens(_cell(row, header, "code tokens")),
            "tags": _split_tokens(_cell(row, header, "tags")),
            "summary_ar": summary_ar,
            "content_clean": content_clean,
            "url": _safe_str(_cell(row, header, "url")),
            "summary_length": summary_length if summary_length is not None else len(summary_ar),
            "content_length": content_length if content_length is not None else len(content_clean),
            "chunk_exists": _to_bool(_cell(row, header, "chunk exists"), default=False),
            "chunk_count": _to_int_or_none(_cell(row, header, "chunk count")) or 0,
            "review_issues": _safe_str(_cell(row, header, "review issues")),
            "is_active": True,
        }
        output_rows.append(record)

    output_rows.sort(
        key=lambda r: (
            _safe_str(r.get("test_name_ar")),
            _safe_str(r.get("title")),
            _safe_str(r.get("id")),
        )
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8") as f:
        for record in output_rows:
            f.write(json.dumps(record, ensure_ascii=False) + "\n")

    return {"written": len(output_rows)}


if __name__ == "__main__":
    stats = build_tests_runtime()
    print(f"INPUT : {INPUT_XLSX_PATH.as_posix()}")
    print(f"SHEET : {INPUT_SHEET_NAME}")
    print(f"OUTPUT: {OUTPUT_JSONL_PATH.as_posix()}")
    print(f"ROWS  : {stats['written']}")
