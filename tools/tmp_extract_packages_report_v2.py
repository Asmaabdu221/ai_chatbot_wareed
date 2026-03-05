import re
from openpyxl import load_workbook

PATH = "app/data/sources/excel/PAKAGE1.xlsx"
SHEET = "Sheet1"

CATEGORY_RANGES = {
    "A1:C1",
    "A47:C49",
    "A103:C105",
    "A108:C108",
    "A111:C112",
    "A116:C118",
}

HEADER_TOKENS = {
    "\u0627\u0633\u0645 \u0627\u0644\u0628\u0627\u0642\u0647",  # اسم الباقه
    "\u0648\u0635\u0641 \u0627\u0644\u0628\u0627\u0642\u0647",  # وصف الباقه
    "\u0633\u0639\u0631 \u0627\u0644\u0628\u0627\u0642\u0647",  # سعر الباقه
    "\u0627\u0633\u0645 \u0627\u0644\u0628\u0627\u0642\u0629",  # اسم الباقة
    "\u0648\u0635\u0641 \u0627\u0644\u0628\u0627\u0642\u0629",  # وصف الباقة
    "\u0633\u0639\u0631 \u0627\u0644\u0628\u0627\u0642\u0629",  # سعر الباقة
}

ARABIC_DIGITS = str.maketrans(
    "\u0660\u0661\u0662\u0663\u0664\u0665\u0666\u0667\u0668\u0669",
    "0123456789",
)


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\n", " ").strip())


def clean_category_v1(v):
    # old behavior for comparison (from prior script)
    return re.sub(r"\|+$", "", norm(v)).strip()


def clean_category_v2(v):
    s = re.sub(r"\|+$", "", norm(v)).strip()
    # Requested split markers
    markers = ["\u0645\u062b\u0644", "\u062b\u0645", "\u060c", ".", ":"]
    # Extra practical marker to match provided example
    markers.append("\u0627\u062c\u0645\u0639")

    split_at = None
    for m in markers:
        idx = s.find(m)
        if idx != -1:
            if split_at is None or idx < split_at:
                split_at = idx
    if split_at is not None:
        s = s[:split_at].strip()
    # Hard cap
    s = s[:40].strip()
    return s


def parse_price_old(price_raw):
    # old behavior: first numeric token
    s = norm(price_raw).translate(ARABIC_DIGITS)
    currency = "SAR" if re.search(r"(\u0631\u064a\u0627\u0644|sar|\ufdfc)", s, flags=re.IGNORECASE) else None
    m = re.search(r"\d+(?:[.,]\d+)?", s)
    if not m:
        return None, currency, s
    try:
        n = float(m.group(0).replace(",", ""))
    except ValueError:
        n = None
    if currency is None and n is not None:
        currency = "SAR"
    return n, currency, s


def parse_price_new(price_raw):
    # Keep raw unchanged; normalize a copy for parsing only.
    s = norm(price_raw).translate(ARABIC_DIGITS)

    currency = "SAR" if re.search(r"(\u0631\u064a\u0627\u0644|sar|\ufdfc)", s, flags=re.IGNORECASE) else None

    # included_count pattern: "<n> تحليل/تحاليل/test/tests"
    included_count = None
    inc = re.search(r"(\d+)\s*(?:\u062a\u062d\u0644\u064a\u0644|\u062a\u062d\u0627\u0644\u064a\u0644|test|tests)\b", s, flags=re.IGNORECASE)
    if inc:
        try:
            included_count = int(inc.group(1))
        except ValueError:
            included_count = None

    # Price rule: if multiple numbers exist, choose LAST number.
    num_matches = list(re.finditer(r"\d+(?:[.,]\d+)?", s))
    if not num_matches:
        return None, currency, included_count, s
    chosen = num_matches[-1]
    try:
        n = float(chosen.group(0).replace(",", ""))
    except ValueError:
        n = None
    if currency is None and n is not None:
        currency = "SAR"
    return n, currency, included_count, s


def extract_rows():
    ws = load_workbook(PATH, data_only=True)[SHEET]

    category_rows = set()
    category_start_old = {}
    category_start_new = {}
    for rg in ws.merged_cells.ranges:
        rg_str = str(rg)
        if rg_str in CATEGORY_RANGES:
            v = ws.cell(rg.min_row, rg.min_col).value
            category_start_old[rg.min_row] = clean_category_v1(v)
            category_start_new[rg.min_row] = clean_category_v2(v)
            for r in range(rg.min_row, rg.max_row + 1):
                category_rows.add(r)

    rows = []
    current_category_old = ""
    current_category_new = ""

    for r in range(1, ws.max_row + 1):
        if r in category_start_old:
            current_category_old = category_start_old[r]
            current_category_new = category_start_new[r]

        a = norm(ws.cell(r, 1).value)
        b = norm(ws.cell(r, 2).value)
        c = norm(ws.cell(r, 3).value)

        if not (a or b or c):
            continue
        if r in category_rows:
            continue
        if a in HEADER_TOKENS or b in HEADER_TOKENS or c in HEADER_TOKENS:
            continue
        if not (a and b and c):
            continue

        old_num, old_cur, old_raw_norm = parse_price_old(c)
        new_num, new_cur, included_count, new_raw_norm = parse_price_new(c)

        rows.append(
            {
                "main_category": current_category_new,
                "package_name": a,
                "description": b,
                "price_raw": c,
                "price_number": new_num,
                "currency": new_cur,
                "included_count": included_count,
                "old_price_number": old_num,
                "old_price_raw_norm": old_raw_norm,
                "new_price_raw_norm": new_raw_norm,
            }
        )

    return rows


def main():
    rows = extract_rows()

    print(f"Input: {PATH}")
    print(f"Sheet: {SHEET}")
    print(f"Total extracted rows after fixes: {len(rows)}")

    print("\nFirst 15 extracted rows:")
    print("main_category | package_name | price_raw | price_number | currency | included_count")
    print("-" * 170)
    for rec in rows[:15]:
        print(
            f"{rec['main_category']} | {rec['package_name']} | "
            f"{rec['price_raw']} | {rec['price_number']} | {rec['currency']} | {rec['included_count']}"
        )

    changed = [r for r in rows if r["old_price_number"] != r["price_number"]]
    print("\nRows where price_number changed (old -> new):")
    if not changed:
        print("(none)")
    else:
        print("package_name | old_price_raw | old_price_number -> new_price_number")
        print("-" * 170)
        for rec in changed:
            print(
                f"{rec['package_name']} | {rec['price_raw']} | "
                f"{rec['old_price_number']} -> {rec['price_number']}"
            )

    cats = sorted({r["main_category"] for r in rows if r["main_category"]})
    print("\nDistinct main_category values (sorted):")
    for c in cats:
        print(f"- {c}")


if __name__ == "__main__":
    main()
