import re
from openpyxl import load_workbook

PATH = "app/data/sources/excel/PAKAGE1.xlsx"
SHEET = "Sheet1"

CATEGORY_RANGES = {
    "A1:C1",
    "A47:C49",
    "A103:C105",
    "A108:C108",
    "A111:C112",
    "A116:C118",
}

HEADER_TOKENS = {
    "\u0627\u0633\u0645 \u0627\u0644\u0628\u0627\u0642\u0647",  # اسم الباقه
    "\u0648\u0635\u0641 \u0627\u0644\u0628\u0627\u0642\u0647",  # وصف الباقه
    "\u0633\u0639\u0631 \u0627\u0644\u0628\u0627\u0642\u0647",  # سعر الباقه
    "\u0627\u0633\u0645 \u0627\u0644\u0628\u0627\u0642\u0629",  # اسم الباقة
    "\u0648\u0635\u0641 \u0627\u0644\u0628\u0627\u0642\u0629",  # وصف الباقة
    "\u0633\u0639\u0631 \u0627\u0644\u0628\u0627\u0642\u0629",  # سعر الباقة
}

ARABIC_DIGITS = str.maketrans(
    "\u0660\u0661\u0662\u0663\u0664\u0665\u0666\u0667\u0668\u0669", "0123456789"
)


def norm(v):
    if v is None:
        return ""
    return re.sub(r"\s+", " ", str(v).replace("\n", " ").strip())


def clean_category(v):
    return re.sub(r"\|+$", "", norm(v)).strip()


def parse_price(price_raw):
    s = norm(price_raw).translate(ARABIC_DIGITS)
    currency = (
        "SAR" if re.search(r"(\u0631\u064a\u0627\u0644|sar|\ufdfc)", s, flags=re.IGNORECASE) else None
    )
    m = re.search(r"\d+(?:[.,]\d+)?", s)
    if not m:
        return None, currency
    try:
        n = float(m.group(0).replace(",", ""))
    except ValueError:
        n = None
    if currency is None and n is not None:
        currency = "SAR"
    return n, currency


def extract_rows():
    ws = load_workbook(PATH, data_only=True)[SHEET]

    category_rows = set()
    category_start = {}
    for rg in ws.merged_cells.ranges:
        rg_str = str(rg)
        if rg_str in CATEGORY_RANGES:
            category_start[rg.min_row] = clean_category(ws.cell(rg.min_row, rg.min_col).value)
            for r in range(rg.min_row, rg.max_row + 1):
                category_rows.add(r)

    rows = []
    current_category = ""
    for r in range(1, ws.max_row + 1):
        if r in category_start:
            current_category = category_start[r]

        a = norm(ws.cell(r, 1).value)
        b = norm(ws.cell(r, 2).value)
        c = norm(ws.cell(r, 3).value)

        if not (a or b or c):
            continue
        if r in category_rows:
            continue
        if a in HEADER_TOKENS or b in HEADER_TOKENS or c in HEADER_TOKENS:
            continue
        if not (a and b and c):
            continue

        price_number, currency = parse_price(c)
        rows.append(
            {
                "main_category": current_category,
                "package_name": a,
                "description": b,
                "price_raw": c,
                "price_number": price_number,
                "currency": currency,
            }
        )

    return rows


def main():
    rows = extract_rows()

    print(f"Input: {PATH}")
    print(f"Sheet: {SHEET}")
    print(f"Total extracted rows: {len(rows)}")

    print("\nFirst 10 extracted rows:")
    print("main_category | package_name | description(truncated) | price_raw | price_number | currency")
    print("-" * 170)
    for rec in rows[:10]:
        d = rec["description"]
        if len(d) > 95:
            d = d[:95] + "..."
        print(
            f"{rec['main_category']} | {rec['package_name']} | {d} | "
            f"{rec['price_raw']} | {rec['price_number']} | {rec['currency']}"
        )

    print("\nDistinct main_category values (sorted):")
    for c in sorted({r["main_category"] for r in rows if r["main_category"]}):
        print(f"- {c}")


if __name__ == "__main__":
    main()

